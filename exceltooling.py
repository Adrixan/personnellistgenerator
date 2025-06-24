from openpyxl import load_workbook
from classes import Teacher, Admin
from mapping import TEACHER_TITLE_PRE, TEACHER_NAME, TEACHER_SURNAME, \
TEACHER_TITLE_POST, TEACHER_KV, TEACHER_KVSTV, TEACHER_SUBJECTS, \
TEACHER_COURSES_1, TEACHER_COURSES_2, TEACHER_COORDINATOR, \
TEACHER_OFFICEHOUR_DAY, TEACHER_OFFICEHOUR_LESSON, TEACHER_AFTERNOON, \
TEACHER_REMARKS, TEACHER_NOIMG, ADMIN_FUNCTION, ADMIN_NAME, ADMIN_SURNAME, ADMIN_TITLE_POST,\
ADMIN_TITLE_PRE

def xstr(s):
    if s is None:
        return ''
    else:
        return str(s)

def booleanize(s):
    if s is None:
        return False
    else:
        return True

def excel_split(inputlist):
    outputlist = []
    if inputlist is not None:
            if ',' in inputlist:
                outputlist = inputlist.split(',')
            else:
                outputlist = [inputlist]
    return outputlist

def to_weekday(shorthand: str):
    if shorthand is not None:
        match shorthand.upper():
            case "MO":
                return "Montag"
            case "DI":
                return "Dienstag"
            case "MI":
                return "Mittwoch"
            case "DO":
                return "Donnerstag"
            case "FR":
                return "Freitag"
            case _:
                return shorthand
    else:
        return ""

def lesson_to_time(lesson: int):
    if lesson is not None:
        match lesson:
            case 1:
                return "7:35-8:25"
            case 2:
                return "8:30-9:20"
            case 3:
                return "9:25-10:15"
            case 4:
                return "10:30-11:20"
            case 5:
                return "11:25-12:15"
            case 6:
                return "12:15-13:05"
            case 7:
                return "13:10-14:00"
            case 8:
                return "14:00-14:50"
            case 9:
                return "14:50-15:40"
            case 10:
                return "15:40-16:30"
            case 11:
                return "16:30-17:20"
            case 12:
                return "17:20-18:10"

def lesson_to_hour_string(lesson_string: str):
    if lesson_string is not None:
        times = []
        if "/" in str(lesson_string):
            lesson_strings = lesson_string.split('/')
            for l in lesson_strings:
                times.append(lesson_to_time(int(l)))
        else:
            times = [lesson_to_time(int(lesson_string))]
        
        counter = 0
        result = ""
        while counter < len(times):
            result = result + times[counter]
            counter = counter + 1
            if counter < len(times):
                result = result + " und "
        return result
    else:
        return "Nach Vereinbarung!"

def convert_excel_to_teachers(infile):
    workbook = load_workbook(filename=infile, read_only=True)
    sheet = workbook["Lehrer"]

    teachers = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[TEACHER_NAME] is None or (row[TEACHER_REMARKS] is not None and "nicht auf die Homepage" in row[TEACHER_REMARKS]):
            continue
        subjects = excel_split(row[TEACHER_SUBJECTS])
        courses1 = excel_split(row[TEACHER_COURSES_1])
        courses2 = excel_split(row[TEACHER_COURSES_2])
        #TODO fix afternoon
        teacher = Teacher(title_pre=xstr(row[TEACHER_TITLE_PRE]), name=xstr(row[TEACHER_NAME]), \
            surname = xstr(row[TEACHER_SURNAME]), title_post=xstr(row[TEACHER_TITLE_POST]), \
            kv=xstr(row[TEACHER_KV]), kvstv=xstr(row[TEACHER_KVSTV]), subjects=subjects, \
            courses_1=courses1, courses_2=courses2, coordinator=xstr(row[TEACHER_COORDINATOR]), \
            officehour_day=to_weekday(row[TEACHER_OFFICEHOUR_DAY]), \
            officehour_lesson=lesson_to_hour_string(row[TEACHER_OFFICEHOUR_LESSON]), \
            afternoon=xstr(row[TEACHER_AFTERNOON]), remarks=row[TEACHER_REMARKS], \
            no_img=booleanize(row[TEACHER_NOIMG])) 
        teachers.append(teacher)
    return teachers

def convert_excel_to_admins(infile):
    workbook = load_workbook(filename=infile, read_only=True)
    sheet = workbook["Verwaltung"]
    admins = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        admin = Admin(function=row[ADMIN_FUNCTION], title_pre=row[ADMIN_TITLE_PRE],\
            name=row[ADMIN_NAME], surname = row[ADMIN_SURNAME])
        admins.append(admin)
    
    return admins